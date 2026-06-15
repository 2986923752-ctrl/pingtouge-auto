"""
平头哥实训平台 - 作业答案自动填入（从 answer_bank.json 批量填入目标账号）

用法:
  python main.py                              # 使用 answer_bank.json 填入
  python main.py --bank answer_bank.json      # 指定答案库
  python main.py --dry-run                    # 仅预览，不实际填入
  python main.py --weeks 4,5                  # 仅处理指定周
  python main.py --headless                   # 无头模式
  python main.py --no-submit                  # 填入但不评测
  python main.py --version                    # 显示版本

运行前请确保:
  1. 已通过 main_build.py 构建好 answer_bank.json
  2. 已准备好 accounts.xlsx（只需「目标账号」列）
  3. 已在 config.py 中配置 DEFAULT_PASSWORD 和 CLASSROOM_URL
"""
from __future__ import annotations

import argparse
import asyncio
import json
import os
import sys
import traceback

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook
from playwright.async_api import Browser, BrowserContext, Page

from config import (
    __version__,
    EXCEL_PATH,
    EXCEL_COL_TARGET_ACCOUNT,
    DEFAULT_PASSWORD,
    ANSWER_BANK_PATH,
    CLASSROOM_URL,
)
from browser import launch_browser, create_context, login, close_context
from extractor import read_combat_levels
from filler import fill_all_answers
from utils import setup_logging, screenshot, wait_for_loading_done


# ─── CLI ───

def parse_args() -> argparse.Namespace:
    """解析命令行参数"""
    p = argparse.ArgumentParser(
        description="平头哥实训平台 - 答案自动填入（从答案库批量填入目标账号）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python main.py                              # 默认运行
  python main.py --dry-run                    # 预览模式
  python main.py --weeks 4,5                  # 仅处理第4、5周
  python main.py --headless --no-submit       # 无头 + 不评测
  python main.py --bank my_bank.json          # 使用自定义答案库
        """,
    )
    p.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    p.add_argument("--bank", type=str, default=ANSWER_BANK_PATH,
                   help="答案库 JSON 文件路径 (默认: answer_bank.json)")
    p.add_argument("--excel", type=str, default=EXCEL_PATH,
                   help="账号 Excel 文件路径 (默认: accounts.xlsx)")
    p.add_argument("--weeks", type=str, default=None,
                   help="指定周，逗号分隔，如 4,5")
    p.add_argument("--classroom", type=str, default=CLASSROOM_URL,
                   help="课堂 URL")
    p.add_argument("--dry-run", action="store_true",
                   help="仅预览不填入")
    p.add_argument("--headless", action="store_true",
                   help="无头模式（后台运行）")
    p.add_argument("--no-submit", action="store_true",
                   help="填入代码但不点击评测")
    return p.parse_args()


# ─── Excel ───

def read_accounts(excel_path: str, logger) -> list[dict]:
    """读取目标账号列表（只需「目标账号」列，密码统一用 DEFAULT_PASSWORD）"""
    wb = load_workbook(excel_path)
    ws = wb.active
    headers: dict[str, int] = {}
    for i, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[cell.value.strip()] = i

    if EXCEL_COL_TARGET_ACCOUNT not in headers:
        logger.error(
            f"缺少「{EXCEL_COL_TARGET_ACCOUNT}」列，当前列: {list(headers.keys())}"
        )
        return []

    accounts: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tgt = row[headers[EXCEL_COL_TARGET_ACCOUNT] - 1]
        if not tgt:
            continue
        accounts.append({
            "target_account": str(tgt).strip(),
            "target_password": DEFAULT_PASSWORD,
        })

    logger.info(f"读取到 {len(accounts)} 个目标账号")
    return accounts


# ─── 页面操作 ───

def get_code_page(ctx: BrowserContext) -> Page | None:
    """从浏览器上下文中找到代码编辑页面"""
    for p in ctx.pages:
        if "/class/code" in p.url:
            return p
    return None


def get_detail_page(ctx: BrowserContext) -> Page | None:
    """从浏览器上下文中找到实训详情页面"""
    for p in ctx.pages:
        if "/class/combat/detail" in p.url:
            return p
    return None


async def get_weeks(page: Page) -> list[str]:
    """扫描课堂页面的所有周标签"""
    return await page.evaluate("""
        () => Array.from(document.querySelectorAll('span'))
            .filter(s => s.textContent.trim().startsWith('Python第') && s.textContent.trim().endsWith('周'))
            .map(s => s.textContent.trim())
    """)


async def scan_experiments(page: Page) -> list[dict]:
    """扫描课堂实验列表，过滤空名（页面未完全加载时的保护）"""
    await wait_for_loading_done(page)
    await page.wait_for_selector(".li-item", timeout=10_000)
    await page.wait_for_timeout(500)
    items = await page.query_selector_all(".li-item")
    result: list[dict] = []
    for item in items:
        name_el = await item.query_selector(".title-con")
        num_el = await item.query_selector(".num-value")
        name = (await name_el.text_content()).strip() if name_el else ""
        num = (await num_el.text_content()).strip() if num_el else "0/0"
        if name:
            result.append({"name": name, "completed": num, "element": item})
    return result


async def click_week(page: Page, week: str) -> None:
    """点击指定周标签，等待加载完成"""
    await page.evaluate(f"""
        () => {{
            const spans = document.querySelectorAll('span');
            for (const s of spans) {{
                if (s.textContent.trim() === '{week}') {{ s.click(); return; }}
            }}
        }}
    """)
    await page.wait_for_timeout(1000)
    await wait_for_loading_done(page)


async def close_extra_pages(ctx: BrowserContext, keep_page: Page) -> None:
    """关闭代码和详情页，回到课堂列表"""
    for p in ctx.pages[:]:
        if p != keep_page and ("/class/code" in p.url or "/class/combat/detail" in p.url):
            await p.close()
    await keep_page.bring_to_front()
    await keep_page.wait_for_timeout(1000)
    await wait_for_loading_done(keep_page)
    await keep_page.wait_for_selector(".li-item", timeout=10_000)
    await keep_page.wait_for_timeout(500)


# ─── 核心: 填入一个实验 ───

async def fill_one_experiment(
    page: Page,
    ctx: BrowserContext,
    exp_name: str,
    src_levels: list[dict],
    logger,
    dry_run: bool,
    no_submit: bool,
) -> bool:
    """填入一个实验的所有未完成关卡"""
    # 重新定位实验并点击「开始实验」
    await wait_for_loading_done(page)
    exps = await scan_experiments(page)
    target = next((e for e in exps if e["name"] == exp_name), None)
    if not target:
        logger.error(f"    未找到实验: {exp_name}")
        return False

    btn = await target["element"].query_selector("button:has-text('开始实验')")
    if not btn:
        logger.error(f"    未找到「开始实验」按钮")
        return False
    await btn.click()
    await page.wait_for_timeout(4000)
    await wait_for_loading_done(page)

    detail = get_detail_page(ctx)
    if not detail:
        logger.error("    未找到详情页")
        return False
    await detail.bring_to_front()
    await detail.wait_for_load_state("networkidle")
    await wait_for_loading_done(detail)
    await detail.wait_for_timeout(2000)

    # 读关卡表
    tgt_levels = await read_combat_levels(detail)
    uncompleted = [i for i, lv in enumerate(tgt_levels) if lv["status"] != "已提交"]
    uncompleted_names = [tgt_levels[i]["name"] for i in uncompleted]
    logger.info(f"    目标未完成: {uncompleted_names}")

    if not uncompleted:
        logger.info("    全部已完成")
        await close_extra_pages(ctx, page)
        return True

    # 点击「进入实验」
    enter = await detail.query_selector("button:has-text('进入实验')")
    if not enter:
        logger.error("    未找到「进入实验」按钮")
        return False
    await enter.click()
    await detail.wait_for_timeout(4000)
    await wait_for_loading_done(detail)

    code_page = get_code_page(ctx) or (detail if "/class/code" in detail.url else None)
    if not code_page:
        logger.error("    未找到代码页面")
        return False
    await code_page.bring_to_front()
    await wait_for_loading_done(code_page)
    await code_page.wait_for_timeout(2000)

    # 匹配答案（跳过源答案库中代码为空的关卡）
    needed: list[dict] = []
    skipped_empty: list[str] = []
    for i in uncompleted:
        if i < len(src_levels):
            code = src_levels[i].get("code", "").strip()
            if len(code) >= 5:
                needed.append(src_levels[i])
            else:
                skipped_empty.append(src_levels[i].get("name", f"#{i+1}"))
    if skipped_empty:
        logger.warning(f"    ⚠ 跳过 {len(skipped_empty)} 个空代码关卡（源答案库不完整）: {skipped_empty}")
    logger.info(f"    填入 {len(needed)} 关 (源共{len(src_levels)}关)")

    if dry_run:
        preview_names = [lv['name'] for lv in needed]
        logger.info(f"    [DRY-RUN] 将填入: {preview_names}")
        await close_extra_pages(ctx, page)
        return True

    if not needed:
        logger.warning("    没有可填入的有效关卡（所有关卡代码均为空）")
        await close_extra_pages(ctx, page)
        return False

    fill_ok = await fill_all_answers(
        code_page, {"type": "code", "code_levels": needed}, logger
    )
    await screenshot(code_page, f"filled_{exp_name}", logger)

    await close_extra_pages(ctx, page)
    return fill_ok


# ─── 主流程 ───

async def main() -> None:
    args = parse_args()

    if args.headless:
        import config
        config.HEADLESS = True

    logger = setup_logging("pingtouge")
    logger.info("=" * 60)
    logger.info(f"平头哥实训平台 - 答案自动填入 v{__version__}")
    logger.info("=" * 60)

    # 读答案库
    if not os.path.exists(args.bank):
        logger.error(f"答案库不存在: {args.bank}")
        logger.info("请先运行: python main_build.py 来构建答案库")
        return
    with open(args.bank, "r", encoding="utf-8") as f:
        bank: dict = json.load(f)
    logger.info(f"答案库: {len(bank.get('experiments', {}))} 个实验")
    logger.info(f"来源账号: {bank.get('source_account', '未知')}")

    # 读账号
    accounts = read_accounts(args.excel, logger)
    if not accounts:
        logger.error("未读取到任何目标账号，请检查 accounts.xlsx")
        return

    if args.dry_run:
        logger.info("模式: DRY-RUN（仅预览）")
    if args.headless:
        logger.info("模式: 无头浏览器")
    if args.no_submit:
        logger.info("模式: 不评测")

    pw, browser = await launch_browser()

    try:
        for i, acc in enumerate(accounts, start=1):
            logger.info(f"\n{'=' * 60}")
            logger.info(f"账号 {i}/{len(accounts)}: {acc['target_account']}")
            logger.info(f"{'=' * 60}")

            ctx = await create_context(browser)
            page = await ctx.new_page()

            try:
                if not await login(
                    page, acc["target_account"], acc["target_password"], logger
                ):
                    logger.error(f"登录失败: {acc['target_account']}")
                    continue

                await page.goto(
                    args.classroom, wait_until="networkidle", timeout=60_000
                )
                await page.wait_for_timeout(3000)
                await wait_for_loading_done(page)
                await page.wait_for_selector(".li-item", timeout=15_000)
                await page.wait_for_timeout(2000)

                weeks = await get_weeks(page)
                if args.weeks:
                    weeks = [f"Python第{w}周" for w in args.weeks.split(",")]
                logger.info(f"扫描周: {weeks}")

                total_filled = 0

                for week in weeks:
                    await click_week(page, week)
                    exps = await scan_experiments(page)

                    # 如果扫出来的实验名都是空的，刷新页面重试
                    if exps and not exps[0]["name"]:
                        logger.warning(f"  {week}: 实验名异常，刷新页面...")
                        await page.reload(wait_until="networkidle")
                        await page.wait_for_timeout(3000)
                        await wait_for_loading_done(page)
                        await click_week(page, week)
                        exps = await scan_experiments(page)

                    logger.info(f"\n{week}: {len(exps)} 个实验")

                    for exp in exps:
                        parts = exp["completed"].split("/")
                        is_done = len(parts) == 2 and parts[0] == parts[1]

                        if is_done:
                            logger.info(f"  跳过 [{exp['name']}] (已完成)")
                            continue

                        if exp["name"] not in bank.get("experiments", {}):
                            logger.warning(f"  无答案 [{exp['name']}]")
                            continue

                        src = bank["experiments"][exp["name"]]
                        src_levels = src.get("code_levels", [])
                        logger.info(
                            f"  填入 [{exp['name']}] (源:{len(src_levels)}关)"
                        )

                        ok = await fill_one_experiment(
                            page, ctx, exp["name"], src_levels,
                            logger, args.dry_run, args.no_submit,
                        )
                        if ok:
                            total_filled += 1

                        # 重新筛选周
                        await click_week(page, week)

                logger.info(
                    f"\n账号 {acc['target_account']} 完成: 填入 {total_filled} 个实验"
                )

            except Exception as e:
                logger.error(f"异常: {e}")
                logger.error(traceback.format_exc())
            finally:
                await close_context(ctx)

    finally:
        await browser.close()
        await pw.stop()
        logger.info("\n全部完成")


if __name__ == "__main__":
    asyncio.run(main())
